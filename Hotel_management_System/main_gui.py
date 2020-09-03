
#######################3
"""
# NAME		:- AYUSH VACHHANI
# SEMESTER	:- 5th
# COURSE	:- PYTHON
# DEPARTMENT	:- ICT
# UNIVERSITY	:- MARWADIUNIVERSITY
"""
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from openpyxl import Workbook
from openpyxl import load_workbook

import datetime 
import mysql.connector

class record:
	def __init__(self,room_no,name,member,mobile,fromdate,todate,room_type):
		self.room_no=room_no
		self.name=name
		self.tmember=member
		self.mobile=mobile
		self.fromdate=fromdate
		self.todate=todate
		self.room_type=room_type
	def __str__(self):
		return "{}#{}#{}#{}#{}#{}#{}#".format(self.room_no,self.name,self.tmember,self.mobile,self.fromdate,self.todate,self.room_type)


	
recordlst=[]
perdayprice=500
					
class Ui_MainWindow(object):
	room_no=0
	selected_date=""
	
	#Connect with database
	def dbconnect(self):
		if(len(recordlst)<1):

			######################################
			#Using notepad File
			
			f1=open("Hotel_management.txt","r")
			recordfromfile=f1.read()
			for rec in recordfromfile.split("\n"):
				rec=rec.split("#")
			
				if len(rec) > 2:
					t1=record(rec[0],rec[1],rec[2],rec[3],rec[4],rec[5],rec[6])
					recordlst.append(t1)

			f1.close()
			

	def sycronizationdata(self):
		f1=open("Hotel_management.txt","w")
		for rec in recordlst:
			print(str(rec),file=f1)
		f1.close()
		
		



	#hide button of all rooms and show field for enter data
	def hideroombtn(self):
		self.hideallroom()
		self.showdatabtn.hide()
		self.calendarWidget.hide()		

		self.roomnolbl.show()
		self.namelbl.show()
		self.t_memberlbl.show()
		self.mobilenolbl.show()
		self.fromdatelbl.show()
		self.todatelbl.show()
		self.insertbtn.show()
		self.updatebtn.show()		
		self.deletebtn.show()
		self.nametxt.show()
		self.membertxt.show()
		self.mobiletxt.show()
		self.fromdatetxt.show()
		self.todatetxt.show()
		self.pricelbl.show()
		self.pricetxt.show()
		self.pricetxt.setDisabled(True)
		self.pricetxt.setText("")
		self.comboBox.show()
		

	def showroombtn(self):
		self.btnroom1.show()
		self.btnroom2.show()
		self.btnroom3.show()
		self.btnroom4.show()
		self.btnroom5.show()
		self.btnroom6.show()
		self.showdatabtn.show()
		self.extractbtn.show()
		self.calendarWidget.show()

		self.roomnolbl.hide()
		self.namelbl.hide()
		self.t_memberlbl.hide()
		self.mobilenolbl.hide()
		self.fromdatelbl.hide()
		self.todatelbl.hide()
		self.insertbtn.hide()
		self.updatebtn.hide()		
		self.deletebtn.hide()
		self.nametxt.hide()
		self.membertxt.hide()
		self.mobiletxt.hide()
		self.fromdatetxt.hide()
		self.todatetxt.hide()
		self.pricelbl.hide()
		self.pricetxt.hide()
		self.comboBox.hide()
		
	def hideallroom(self):
		self.btnroom1.hide()
		self.btnroom2.hide()
		self.btnroom3.hide()
		self.btnroom4.hide()
		self.btnroom5.hide()
		self.btnroom6.hide()
		self.extractbtn.hide()
		


	def showdata(self):
		self.showroombtn()
		str1=str(self.calendarWidget.selectedDate())
		date=str1[19 :len(str1)-1 ]
		date=date.split(",")
		self.selected_date="{}/{}/{}".format(date[2],date[1],date[0])
		self.dbconnect()
		self.changebtncolor()		
	

	def setdefault(self):
		self.roomnolbl.setText("Room-"+str(self.room_no))
		temp=0
		for rec in recordlst:
			rec=str(rec)
			rec=rec.split("#")
			if(rec[0] == str(self.room_no) and self.date_in_between(rec[4],rec[5],self.selected_date)):
				
				self.insertbtn.setDisabled(True)
				self.updatebtn.setDisabled(False)
				self.deletebtn.setDisabled(False)
				self.nametxt.setText(rec[1])
				self.membertxt.setText(rec[2])
				self.mobiletxt.setText(rec[3])
				self.fromdatetxt.setText(rec[4])
				self.fromdatetxt.setDisabled(False)
				self.todatetxt.setText(rec[5])
				
				if rec[6]=="AC":
					self.comboBox.setCurrentIndex(0)
				else:
					self.comboBox.setCurrentIndex(1)

				fromdate=str(rec[4]).split("/")
				todate=str(rec[5]).split("/")

				d1=datetime.datetime(int(fromdate[2]),int(fromdate[1]),int(fromdate[0]))
				d2=datetime.datetime(int(todate[2]),int(todate[1]),int(todate[0]))

				t1=str(d2-d1);
				t1=t1.split(" ")
				if(rec[6]=="AC"):
					self.pricetxt.setText(str(int(t1[0])*perdayprice*2))
				else:
					self.pricetxt.setText(str(int(t1[0])*perdayprice))
				temp=1
		if(temp==0):
			self.nametxt.setText("")
			self.membertxt.setText("")
			self.mobiletxt.setText("")
			self.fromdatetxt.setText(self.selected_date)
			self.fromdatetxt.setDisabled(True)
			self.todatetxt.setText("")
			self.insertbtn.setDisabled(False)
			self.updatebtn.setDisabled(True)
			self.deletebtn.setDisabled(True)

	def handlerecord(self,room):
		self.room_no=room
		self.hideroombtn()
		self.setdefault()

	def date_in_between(self,fromdate,todate,selected_date):		
		fromdate=fromdate.split("/")
		todate=todate.split("/")
		selected_date=selected_date.split("/")
	
		d1=datetime.datetime(int(fromdate[2]),int(fromdate[1]),int(fromdate[0]))
		d2=datetime.datetime(int(selected_date[2]),int(selected_date[1]),int(selected_date[0]))
		d3=datetime.datetime(int(todate[2]),int(todate[1]),int(todate[0]))
	
		if (d1<=d2<d3):
			return 1
		else:
			return 0
					

	def changebtncolor(self):
		self.btnroom1.setStyleSheet("background-color : yellow")
		self.btnroom2.setStyleSheet("background-color : yellow") 
		self.btnroom3.setStyleSheet("background-color : yellow") 
		self.btnroom4.setStyleSheet("background-color : yellow") 
		self.btnroom5.setStyleSheet("background-color : yellow") 
		self.btnroom6.setStyleSheet("background-color : yellow")  
		
		for rec in recordlst:
			rec=str(rec)
			rec=rec.split("#")
			
			if(rec[0]=="1" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom1.setStyleSheet("background-color : red") 
			elif(rec[0]=="2" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom2.setStyleSheet("background-color : red") 
			elif(rec[0]=="3" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom3.setStyleSheet("background-color : red") 
			elif(rec[0]=="4" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom4.setStyleSheet("background-color : red") 
			elif(rec[0]=="5" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom5.setStyleSheet("background-color : red") 
			elif(rec[0]=="6" and self.date_in_between(rec[4],rec[5],self.selected_date)):
				self.btnroom6.setStyleSheet("background-color : red") 



	def checkinputs(self,name,member,mobileno,fromdate,todate):
		name=name.split(" ")

		if (len(name)!=2 or not(name[0].isalpha())  or not(name[1].isalpha())):
			raise ValueError("Name is Not in proper format")
			
		if (not(str(member).isdigit())):
			raise ValueError("Member is Not in proper format")
			
		if (not(str(mobileno).isdigit()) or len(str(mobileno))!=10):
			raise ValueError("Mobile no. is Not in proper format")
			
		fromdate=fromdate.split("/")
		todate=todate.split("/")
		
		d1=datetime.datetime(int(fromdate[2]),int(fromdate[1]),int(fromdate[0]))
		d2=datetime.datetime(int(todate[2]),int(todate[1]),int(todate[0]))
	
		if (d1>=d2):
			raise ValueError("Date is not in proper format")
			return 0
		
		
		for rec in recordlst:
			rec=str(rec)
			rec=rec.split("#")
			fdate=rec[4]
			ldate=rec[5]
			
			fdate=fdate.split("/")
			ldate=ldate.split("/")

			d3=datetime.datetime(int(fdate[2]),int(fdate[1]),int(fdate[0]))
			d4=datetime.datetime(int(ldate[2]),int(ldate[1]),int(ldate[0]))
			if(d3<d2<d4 and self.room_no==int(rec[0])):
				raise ValueError("This room is not Empty between this dates")
			if(d3<d1<d4 and self.room_no==int(rec[0])):
				raise ValueError("This room is not Empty between this dates")
			if(d1<d3<d2 and self.room_no==int(rec[0])):
				raise ValueError("This room is not Empty between this dates")
			if(d1<d3<d2 and self.room_no==int(rec[0])):
				raise ValueError("This room is not Empty between this dates")
		return 1

	def insertdata(self):
		
		name=self.nametxt.text()
		member=self.membertxt.text()
		mobileno=self.mobiletxt.text()
		fromdate=self.fromdatetxt.text()
		todate=self.todatetxt.text()
		room_type=self.comboBox.currentText()
		try:
			true=self.checkinputs(name,member,mobileno,fromdate,todate)
			r1=record(self.room_no,name,member,mobileno,fromdate,todate,room_type)
		
			recordlst.append(r1)
			
			
			self.sycronizationdata()
			self.showroombtn()
			
			self.changebtncolor()
		except BaseException as e:
			msg=QMessageBox()
			msg.setWindowTitle("Error")
			msg.setText("{}".format(e))
			msg.setIcon(QMessageBox.Critical)
			msg.exec_()
	
	def updatedata(self):
		len1=len(recordlst)
		temp=0
		try:
			
			for i in range(len1):
				i1=recordlst.pop()
				s1=str(i1)
				s1=s1.split("#")
				if int(s1[0])==self.room_no  and self.date_in_between(s1[4],s1[5],self.selected_date):
					name=self.nametxt.text()
					member=self.membertxt.text()
					mobileno=self.mobiletxt.text()
					fromdate=self.fromdatetxt.text()
					todate=self.todatetxt.text()
					room_type=self.comboBox.currentText()
					
					if(self.checkinputs(name,member,mobileno,fromdate,todate)):
						i1=record(self.room_no,name,member,mobileno,fromdate,todate,room_type)
						temp=1

				recordlst.insert(0,i1)
			self.sycronizationdata()
			self.showroombtn()
			self.changebtncolor()	
		except BaseException as e:
			recordlst.insert(0,i1)
			msg=QMessageBox()
			msg.setWindowTitle("Error")
			msg.setText("{}".format(e))
			msg.setIcon(QMessageBox.Critical)
			msg.exec_()
	


	def deletedata(self):
		len1=len(recordlst)
		for i in range(len1):
			i1=recordlst.pop()
			s1=str(i1)
			s1=s1.split("#")
			if s1[0]!=str(self.room_no) or not(self.date_in_between(s1[4],s1[5],self.selected_date)):
				
				recordlst.insert(0,i1)
						

		self.sycronizationdata()
		self.showroombtn()
		self.changebtncolor()

	def extractData(self):
		mywb = Workbook()  #Create New Workbook
		mysheet = mywb.active  #select first sheet
		
		mysheet.cell(1,1,"sr.no")
		mysheet.cell(1,2,"Room No.")
		mysheet.cell(1,3,"Name")
		mysheet.cell(1,4,"Total Member")
		mysheet.cell(1,5,"Mobile No")
		mysheet.cell(1,6,"From date")
		mysheet.cell(1,7,"To data")
		mysheet.cell(1,8,"Room Type")
		sr_no=1
		for rec in recordlst:
			rec=str(rec)		
			rec=rec.split("#")
			mysheet.cell(sr_no+1,1,sr_no)
			mysheet.cell(sr_no+1,2,rec[0])
			mysheet.cell(sr_no+1,3,rec[1])
			mysheet.cell(sr_no+1,4,rec[2])
			mysheet.cell(sr_no+1,5,rec[3])
			mysheet.cell(sr_no+1,6,rec[4])
			mysheet.cell(sr_no+1,7,rec[5])
			mysheet.cell(sr_no+1,8,rec[6])
			sr_no+=1
		mywb.save("Hotel-Data.xlsx")



	def setupUi(self, MainWindow):
		MainWindow.setObjectName("MainWindow")
		MainWindow.resize(642, 662)
		
		self.centralwidget = QtWidgets.QWidget(MainWindow)
		self.centralwidget.setObjectName("centralwidget")
		#self.centralwidget.setStyleSheet("background-color: blue;") 		

		###################
		self.btnroom1 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom1.setGeometry(QtCore.QRect(20, 200, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom1.setFont(font)
		self.btnroom1.setObjectName("btnroom1")
		
		self.btnroom2 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom2.setGeometry(QtCore.QRect(230, 200, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom2.setFont(font)
		self.btnroom2.setObjectName("btnroom2")

		self.btnroom3 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom3.setGeometry(QtCore.QRect(440, 200, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom3.setFont(font)
		self.btnroom3.setObjectName("btnroom3")
				

		self.btnroom4 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom4.setGeometry(QtCore.QRect(20, 420, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom4.setFont(font)
		self.btnroom4.setObjectName("btnroom4")
		
		self.btnroom5 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom5.setGeometry(QtCore.QRect(230, 420, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom5.setFont(font)
		self.btnroom5.setObjectName("btnroom5")
		
		self.btnroom6 = QtWidgets.QPushButton(self.centralwidget)
		self.btnroom6.setGeometry(QtCore.QRect(440, 420, 181, 201))
		font = QtGui.QFont()
		font.setPointSize(20)
		self.btnroom6.setFont(font)
		self.btnroom6.setObjectName("btnroom6")

		self.calendarWidget = QtWidgets.QCalendarWidget(self.centralwidget)
		self.calendarWidget.setGeometry(QtCore.QRect(20, 0, 312, 183))
		self.calendarWidget.setObjectName("calendarWidget")

		self.showdatabtn = QtWidgets.QPushButton(self.centralwidget)
		self.showdatabtn.setGeometry(QtCore.QRect(380, 30, 191, 61))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.showdatabtn.setFont(font)
		self.showdatabtn.setObjectName("showdatabtn")
		
		self.extractbtn = QtWidgets.QPushButton(self.centralwidget)
		self.extractbtn.setGeometry(QtCore.QRect(380, 110, 191, 61))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.extractbtn.setFont(font)
		self.extractbtn.setObjectName("extractbtn")


		#####################				


		self.roomnolbl = QtWidgets.QLabel(MainWindow)
		self.roomnolbl.setGeometry(QtCore.QRect(230, 0, 141, 51))
		font = QtGui.QFont()
		font.setPointSize(25)
		self.roomnolbl.setFont(font)
		self.roomnolbl.setObjectName("roomnolbl")

		self.namelbl = QtWidgets.QLabel(MainWindow)
		self.namelbl.setGeometry(QtCore.QRect(40, 100, 91, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.namelbl.setFont(font)
		self.namelbl.setObjectName("namelbl")
		
		self.t_memberlbl = QtWidgets.QLabel(MainWindow)
		self.t_memberlbl.setGeometry(QtCore.QRect(40, 170, 141, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.t_memberlbl.setFont(font)
		self.t_memberlbl.setObjectName("t_memberlbl")

		self.mobilenolbl = QtWidgets.QLabel(MainWindow)
		self.mobilenolbl.setGeometry(QtCore.QRect(40, 240, 141, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.mobilenolbl.setFont(font)
		self.mobilenolbl.setObjectName("mobilenolbl")		
		
		self.fromdatelbl = QtWidgets.QLabel(MainWindow)
		self.fromdatelbl.setGeometry(QtCore.QRect(40, 310, 141, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.fromdatelbl.setFont(font)
		self.fromdatelbl.setObjectName("fromdatelbl")
		
		self.todatelbl = QtWidgets.QLabel(MainWindow)
		self.todatelbl.setGeometry(QtCore.QRect(400, 310, 141, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.todatelbl.setFont(font)
		self.todatelbl.setObjectName("todatelbl")

		self.insertbtn = QtWidgets.QPushButton(MainWindow)
		self.insertbtn.setGeometry(QtCore.QRect(40, 430, 141, 51))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.insertbtn.setFont(font)
		self.insertbtn.setObjectName("insertbtn")

		self.updatebtn = QtWidgets.QPushButton(MainWindow)
		self.updatebtn.setGeometry(QtCore.QRect(230, 430, 141, 51))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.updatebtn.setFont(font)
		self.updatebtn.setObjectName("updatebtn")

		self.deletebtn = QtWidgets.QPushButton(MainWindow)
		self.deletebtn.setGeometry(QtCore.QRect(420, 430, 141, 51))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.deletebtn.setFont(font)
		self.deletebtn.setObjectName("deletebtn")
		
		self.nametxt = QtWidgets.QLineEdit(MainWindow)
		self.nametxt.setGeometry(QtCore.QRect(220, 100, 321, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.nametxt.setFont(font)
		self.nametxt.setText("")
		self.nametxt.setObjectName("nametxt")
	
		self.membertxt = QtWidgets.QLineEdit(MainWindow)
		self.membertxt.setGeometry(QtCore.QRect(220, 170, 151, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.membertxt.setFont(font)
		self.membertxt.setText("")
		self.membertxt.setObjectName("membertxt")

			
		self.mobiletxt = QtWidgets.QLineEdit(MainWindow)
		self.mobiletxt.setGeometry(QtCore.QRect(220, 240, 321, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.mobiletxt.setFont(font)
		self.mobiletxt.setText("")
		self.mobiletxt.setObjectName("mobiletxt")
	
		self.fromdatetxt = QtWidgets.QLineEdit(MainWindow)
		self.fromdatetxt.setGeometry(QtCore.QRect(220, 310, 151, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.fromdatetxt.setFont(font)
		self.fromdatetxt.setText("")
		self.fromdatetxt.setObjectName("fromdatetxt")
	
		self.todatetxt = QtWidgets.QLineEdit(MainWindow)
		self.todatetxt.setGeometry(QtCore.QRect(450, 310, 151, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.todatetxt.setFont(font)
		self.todatetxt.setText("")
		self.todatetxt.setObjectName("fromdatetxt")


		self.comboBox = QtWidgets.QComboBox(MainWindow)
		self.comboBox.setGeometry(QtCore.QRect(430, 380, 121, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.comboBox.setFont(font)
		self.comboBox.setObjectName("comboBox")
		self.comboBox.addItem("")
		self.comboBox.addItem("")

		self.pricelbl = QtWidgets.QLabel(MainWindow)
		self.pricelbl.setGeometry(QtCore.QRect(40, 380, 141, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.pricelbl.setFont(font)
		self.pricelbl.setObjectName("pricelbl")
        
		self.pricetxt = QtWidgets.QLineEdit(MainWindow)
		self.pricetxt.setGeometry(QtCore.QRect(220, 380, 151, 31))
		font = QtGui.QFont()
		font.setPointSize(15)
		self.pricetxt.setFont(font)
		self.pricetxt.setText("")
		self.pricetxt.setObjectName("pricetxt")








		########################

		MainWindow.setCentralWidget(self.centralwidget)
		self.menubar = QtWidgets.QMenuBar(MainWindow)
		self.menubar.setGeometry(QtCore.QRect(0, 0, 642, 21))
		self.menubar.setObjectName("menubar")
		MainWindow.setMenuBar(self.menubar)
		self.statusbar = QtWidgets.QStatusBar(MainWindow)
		self.statusbar.setObjectName("statusbar")
		MainWindow.setStatusBar(self.statusbar)
		

		self.showroombtn()
		self.hideallroom()
	
				

		self.insertbtn.clicked.connect(self.insertdata)
		self.updatebtn.clicked.connect(self.updatedata)
		self.deletebtn.clicked.connect(self.deletedata)	

		self.btnroom1.clicked.connect(lambda: self.handlerecord(1))
		self.btnroom2.clicked.connect(lambda: self.handlerecord(2))
		self.btnroom3.clicked.connect(lambda: self.handlerecord(3))
		self.btnroom4.clicked.connect(lambda: self.handlerecord(4))
		self.btnroom5.clicked.connect(lambda: self.handlerecord(5))
		self.btnroom6.clicked.connect(lambda: self.handlerecord(6))
		
		self.showdatabtn.clicked.connect(self.showdata)	
		self.extractbtn.clicked.connect(self.extractData)						

		self.retranslateUi(MainWindow)
		QtCore.QMetaObject.connectSlotsByName(MainWindow)

	def retranslateUi(self, MainWindow):
		_translate = QtCore.QCoreApplication.translate
		MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
		self.btnroom1.setText(_translate("MainWindow", "Room-1"))
		self.btnroom2.setText(_translate("MainWindow", "Room-2"))
		self.btnroom4.setText(_translate("MainWindow", "Room-4"))
		self.btnroom5.setText(_translate("MainWindow", "Room-5"))
		self.btnroom3.setText(_translate("MainWindow", "Room-3"))
		self.btnroom6.setText(_translate("MainWindow", "Room-6"))
		self.showdatabtn.setText(_translate("MainWindow", "Show Data"))
		self.extractbtn.setText(_translate("MainWindow", "Extract Data"))
			
		#################
		
		self.roomnolbl.setText(_translate("MainWindow", "Room No."))
		self.namelbl.setText(_translate("MainWindow", "Name-"))
		self.t_memberlbl.setText(_translate("MainWindow", "Total Member-"))
		self.mobilenolbl.setText(_translate("MainWindow", "Mobile No.-"))
		self.fromdatelbl.setText(_translate("MainWindow", "From -"))
		self.insertbtn.setText(_translate("MainWindow", "INSERT"))
		self.updatebtn.setText(_translate("MainWindow", "UPDATE"))
		self.deletebtn.setText(_translate("MainWindow", "DELETE"))
		self.todatelbl.setText(_translate("MainWindow", "To -"))
		self.pricelbl.setText(_translate("MainWindow", "Price -"))
		self.comboBox.setItemText(0, _translate("MainWindow", "AC"))
		self.comboBox.setItemText(1, _translate("MainWindow", "Non AC"))

		
	

"""
import sys
app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()
sys.exit(app.exec_())
"""



